# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from functools import reduce
import json

wb = load_workbook(filename='1.xlsx')
ws = wb.active
super_dict = {}

# replace all none cells with upper cell, by column
for col in ws.iter_cols():
    tmp = ''
    for cell in col:
        if not cell.value:
            cell.value = tmp
        else:
            tmp = cell.value

# put everything in one dict
for row in ws.iter_rows():  # full table or use min_row=3, max_col=4, max_row=10 to debug on few rows
    #parse cell row2list
    c = []
    for cell in row:
        c.append(cell.value)
    #list2dict
    d = reduce(lambda x, y: {'title': y, 'children': [x]}, reversed(c), {})
    #dict2superdicts
    for key, value in d.items():
        super_dict.setdefault(key, []).append(value)

#print(super_dict)
print(json.dumps(super_dict,sort_keys=True, indent=4, ensure_ascii=False, separators=(',', ': ')))